"""Salary `.xlsx` parser — extract-and-return (ADR-0010 pattern, ADR-0014).

hr-ai parses the workbook and RETURNS structured rows; hr-backend writes
`salary_tables` / `salary_table_rows` / `convenio_job_categories`. hr-ai writes
NO salary DB rows.

Handles the real messiness seen across COEAS Andalucía / Deporte Cantabria /
COEAS Estatal:
- ignore junk/notes sheets (no salary-grid header, or tiny);
- the header row is NOT row 1 (scan for it);
- map cryptic columns via a header-synonym map (per-format maps converge on one
  synonym set);
- ONE workbook → MANY year tables (e.g. `smi 26` + `smi 25`), one per sheet
  (ADR-0014 / plan §9 Q3).

Canonical 14/12 mapping (catch 3): base_salary_monthly = gross_annual / 14 and
num_payments = 14 (Spanish 12-monthly-plus-2-extras norm). The /12 figure and
ALL original columns (SB, COMP, Comp. SMI, the 14 & 12 figures, totals…) are
kept verbatim in raw_values. Documented in data-model.md §6.
"""

from __future__ import annotations

import io
import re
import unicodedata

import openpyxl

_NUM_PAYMENTS = 14


def _norm(text) -> str:
    if text is None:
        return ""
    s = str(text).replace("\n", " ")
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s.strip(" .:·-")


# Header-synonym sets (normalized) for the TYPED columns we extract. These
# converge across the real per-format files (COEAS Andalucía/Estatal, Deporte
# Cantabria/Estatal, Agencias, Limpieza Navarra, …) — plan §9 Q2.
_GROSS = {"total", "total anual", "bruto anual", "bruto ano", "bruto año", "importe anual", "salario anual"}
_HOURLY = {
    "€/hora", "e/hora", "euro/hora", "euros/hora", "hora", "precio hora", "precio/hora",
    "coste hora", "/hora", "bruto/hora", "bruto hora", "salario hora",
}
_EXTRA = {"pagas extra", "paga extra", "pagas extras", "paga extras"}
_NIGHT = {"plus nocturno", "nocturnidad", "plus noche", "nocturno", "plus nocturnidad",
          "plus hora nocturna", "plus hora noctur", "hora nocturna"}

# Raw money-column markers (kept verbatim in raw_values, used to anchor where the
# numeric grid starts — everything LEFT of the first money header is a label).
_RAW_MONEY = {
    "sb", "sb anual", "comp", "comp.", "comp smi", "comp. smi", "comp smi / ano",
    "comp smi / mes", "14", "12", "bruto mes", "bruto/mes 14 pagas", "bruto/mes 12 pagas",
    "salario base", "dedica", "pc", "paga 16", "p.p.paga extra", "plus transporte",
    "plus tpte/dia", "plus tpte/día", "5% mejora sedena", "1,2,3,5 quinquenio",
    "4 quinquenio", "quinquenio", "antiguedad",
}

# A header cell that anchors the start of the numeric grid (typed OR raw money).
_MONEY_HEADERS = _GROSS | _HOURLY | _EXTRA | _NIGHT | _RAW_MONEY

# Tokens that mark a row as a salary-grid HEADER.
_HEADER_MARKERS = _MONEY_HEADERS


def _is_number(v) -> bool:
    if isinstance(v, (int, float)):
        return True
    if v is None:
        return False
    s = str(v).strip().replace(".", "").replace(",", "")
    return s.isdigit()


def _to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    # Spanish decimal-comma normalization (1.652,13 → 1652.13), as registry import.
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _looks_like_hours_header(label: str) -> bool:
    # The "1742" column (a bare annual-hours number as header) holds €/hora.
    return bool(re.fullmatch(r"\d{3,4}", label))


def _year_from_sheet_name(name: str):
    m = re.search(r"(19|20)\d{2}", name)
    if m:
        return int(m.group(0))
    m = re.search(r"\b(\d{2})\b", name)  # "smi 26" → 2026
    if m:
        return 2000 + int(m.group(1))
    return None


def _find_header_row(rows: list[list]) -> int | None:
    best_idx, best_score = None, 0
    for i, row in enumerate(rows[:12]):
        score = 0
        for cell in row:
            label = _norm(cell)
            if not label:
                continue
            if label in _HEADER_MARKERS or _looks_like_hours_header(label) or label in _GROSS:
                score += 1
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx if best_score >= 2 else None


def _column_field_map(header: list[str]) -> dict:
    """col index → typed field name (gross_annual / hourly_rate / extra_pay /
    night_plus). Unmapped numeric columns stay raw_values only."""
    mapping = {}
    for idx, raw in enumerate(header):
        label = _norm(raw)
        if not label:
            continue
        if label in _GROSS and "gross_annual" not in mapping.values():
            mapping[idx] = "gross_annual"
        elif label in _HOURLY or _looks_like_hours_header(label):
            mapping.setdefault(idx, "hourly_rate")
        elif label in _EXTRA:
            mapping.setdefault(idx, "extra_pay")
        elif label in _NIGHT:
            mapping.setdefault(idx, "night_plus")
    return mapping


def _parse_sheet(name: str, rows: list[list]) -> dict | None:
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return None
    width = _width(rows)
    header = [_norm_or_index(rows[header_idx], i) for i in range(width)]
    field_map = _column_field_map(rows[header_idx])
    data = rows[header_idx + 1:]

    # Label columns = everything LEFT of the first money-header column (the
    # numeric grid). Robust across formats whose label is a code that *looks*
    # numeric (Cantabria "3.1") — header position, not value type, decides.
    money_cols = [c for c in range(width) if header[c] in _MONEY_HEADERS or _looks_like_hours_header(header[c])]
    if money_cols:
        first_money = min(money_cols)
    else:
        # Fallback: leftmost column whose data is mostly numeric.
        numeric_cols = []
        for c in range(width):
            nonempty = [r[c] for r in data if c < len(r) and r[c] not in (None, "")]
            if nonempty and sum(_is_number(v) for v in nonempty) / len(nonempty) > 0.5:
                numeric_cols.append(c)
        first_money = min(numeric_cols) if numeric_cols else width
    label_cols = list(range(first_money))

    out_rows = []
    for r in data:
        # A data row must carry at least one numeric value in the grid.
        if not any(c < len(r) and _is_number(r[c]) for c in range(first_money, width)):
            continue
        # Label columns → group_code (leftmost) + job_category_name (rightmost),
        # skipping genuinely empty cells (never the literal string "None").
        labels = []
        for c in label_cols:
            if c < len(r) and r[c] is not None:
                s = re.sub(r"\s+", " ", str(r[c])).strip()  # collapse embedded newlines
                s = s.strip("'\u2019\u2018\"`").strip()  # strip wrapping quotes/apostrophes (e.g. "2.1'" → "2.1")
                if s:
                    labels.append(s)
        if not labels:
            continue
        job_category_name = labels[-1]
        group_code = labels[0] if len(labels) >= 2 else (
            labels[0] if re.match(r"^\d+(\.\d+)?'?$", labels[0]) else None
        )

        gross = None
        hourly = extra = night = None
        raw_values = {}
        for c in range(width):
            if c in label_cols or c >= len(r):
                continue
            val = r[c]
            if val in (None, ""):
                continue
            label = header[c] if header[c] else f"col{c}"
            raw_values[label] = val if not isinstance(val, float) else round(val, 6)
            field = field_map.get(c)
            if field == "gross_annual":
                gross = _to_float(val)
            elif field == "hourly_rate":
                hourly = _to_float(val)
            elif field == "extra_pay":
                extra = _to_float(val)
            elif field == "night_plus":
                night = _to_float(val)

        base_monthly = round(gross / _NUM_PAYMENTS, 2) if gross is not None else None
        out_rows.append(
            {
                "job_category_name": job_category_name,
                "group_code": group_code,
                "gross_annual": round(gross, 2) if gross is not None else None,
                "base_salary_monthly": base_monthly,
                "extra_pay": round(extra, 2) if extra is not None else None,
                "num_payments": _NUM_PAYMENTS if gross is not None else None,
                "hourly_rate": round(hourly, 4) if hourly is not None else None,
                "night_plus": round(night, 2) if night is not None else None,
                "raw_values": raw_values,
            }
        )

    if not out_rows:
        return None
    return {
        "sheet": name,
        "year": _year_from_sheet_name(name),
        "validity_start": None,
        "validity_end": None,
        "rows": out_rows,
    }


def _width(rows: list[list]) -> int:
    return max((len(r) for r in rows), default=0)


def _norm_or_index(row: list, i: int) -> str:
    return _norm(row[i]) if i < len(row) else ""


def parse_salary_xlsx(xlsx_bytes: bytes) -> dict:
    """Parse a salary workbook → {tables: [...], warnings: [...]}.

    One table per salary sheet (multi-year supported). Junk/notes sheets that
    have no salary-grid header are skipped and reported in warnings.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    tables, warnings = [], []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows or _width(rows) < 2:
            warnings.append(f"sheet '{name}' skipped (empty/too small)")
            continue
        parsed = _parse_sheet(name, rows)
        if parsed is None:
            warnings.append(f"sheet '{name}' skipped (no salary-grid header found)")
            continue
        hdr = _find_header_row(rows)
        warnings.append(f"sheet '{name}': header on row {hdr}, {len(parsed['rows'])} category rows, year {parsed['year']}")
        tables.append(parsed)
    return {"tables": tables, "warnings": warnings}
